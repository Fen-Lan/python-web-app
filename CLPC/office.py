# from openpyxl import Workbook
import os.path
import shutil
import subprocess
import time
import traceback
from time import sleep

from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook as OpenpyxlWorkbook
from typing import Union, List,Any
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries, coordinate_to_tuple
import copy, sys


from CLPC.framework import FRAME, class_func_tracker
from CLPC.tool import TOOL
from CLPC.framework import FUNC_USAGE_TRACKER

@FUNC_USAGE_TRACKER
class excel(OpenpyxlWorkbook):
    """
    todo
    1. 服务器安装libreoffice
    2. 打开所有excel文件时，用libreoffice打开文件，并另存为，可以排除一些文件无法打开的问题，但是会丢失格式
    3. 在转换成图片之前，也用libreoffice重新打开一下文件，保证公式被计算。
    在使用脚本之前，需要确保 libreoffice 已经启动
    """
    @staticmethod
    def check_excel_file_format(excel_file):
        """
        识别当前 excel 文件的实际格式，通过修改后缀方式无法修改其实际格式

        """
        import xlrd
        name, excel_type = os.path.splitext(os.path.basename(excel_file))
        try:
            bk = xlrd.open_workbook(excel_file, formatting_info=True)
            result = "xls"
            bk.release_resources()
        except NotImplementedError:
            # print("此文件实际格式为 xlsx")
            result = "xlsx"
        except:
            print("当前文件可能损坏无法读取")
            # print(traceback.format_exc())
            return True  # 读取异常时，按照初始文件后缀格式处理

        if excel_type == '.xls' and result == 'xls':
            print("实际格式与后缀一致，为 xls 格式，按 xls 格式处理")
            return True
        if excel_type == '.xlsx' and result == 'xlsx':
            print("实际格式与后缀一致，为 xlsx 格式，按 xlsx 格式处理")
            return True
        if excel_type == '.xls' and result == 'xlsx':
            print("实际格式与后缀不一致，实际为 xlsx 格式，按 xlsx 格式处理")
            return False
        if excel_type == '.xlsx' and result == 'xls':
            print("实际格式与后缀不一致，实际为 xls 格式，按 xls 格式处理")
            return False

    @staticmethod
    def xls2xlsx(xls_file, xlsx_file):
        """
        讲xls文件转换为xlsx文件，然后可以用组件处理
        """
        from xls2xlsx import XLS2XLSX
        x2x = XLS2XLSX(xls_file)
        try:
            x2x.to_xlsx(xlsx_file)
        except Exception as e:
            print("xls转换异常：", e)
            raise Exception("xls文件处理异常，建议手动存为xlsx格式后处理")
        import os
        # os.system("soffice --headless --convert-to xlsx {}".format(xls_file))

    @staticmethod
    def convert_xls_file(xls_file:str):
        """
        使用第三方软件转换xls文件格式到xlsx格式，windows下使用wps，mac、linux下使用libreoffice
        输出文件名为将xls替换为xlsx
        """
        sysinfo = FRAME.get_operating_system()
        exec_env = FRAME.get_env_arg("RPA_ENV", "")
        result_file_name = xls_file.replace(".xls", '.xlsx')
        if exec_env == "prod" or exec_env == 'test': # 服务器环境执行
            # 服务器环境待适配
            convert_result = subprocess.run(["/opt/libreoffice25.2/program/python", "CLPC/uno_script.py", "convert", xls_file],
                                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if convert_result.returncode != 0:
                print("文件转换过程日志：", convert_result.stdout.decode())
                raise Exception("LibreOffice转换失败: {}".format(convert_result.stderr.decode()))
            else:
                print("LibreOffice转换成功")
                return result_file_name
        else:
            if sysinfo == "MacOS" :
                convert_result = subprocess.run(["/Applications/LibreOffice.app/Contents/Resources/python", "CLPC/uno_script.py", "convert", xls_file],
                                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if convert_result.returncode != 0:
                    print("文件转换过程日志：", convert_result.stdout.decode())
                    raise Exception("LibreOffice转换失败: {}".format(convert_result.stderr.decode()))
                else:
                    print("LibreOffice转换成功")
                return result_file_name

            elif sysinfo == "Linux":
                pass
                # 信创环境后续适配
            elif sysinfo == "Windows":
                from win32com.client import Dispatch
                try:
                    xlApp = Dispatch("Ket.Application")  # 使用wps， 如果使用office，需要改为 "Excel.Application"
                except:
                    print("请确认是否安装WPS软件")
                try:
                    xlApp.Visible = False
                    xlApp.DisplayAlerts = False
                    dir_name = os.path.dirname(xls_file)
                    if "\\" in xls_file or  "/" in xls_file or dir_name != "":
                        pass
                    else:
                        main_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
                        os.path.basename(xls_file)
                        xls_file = os.path.join(main_dir, xls_file)
                    xlApp.Workbooks.Open(xls_file)
                    result_xlsx_file = os.path.splitext(xls_file)[0] + ".xlsx"
                    xlApp.ActiveWorkbook.SaveAs(result_xlsx_file, FileFormat=51)
                    xlApp.ActiveWorkbook.Close()
                    xlApp.Quit()
                    return result_file_name

                except:
                    raise Exception("请确认文件存在，或者文件没有被其他程序占用")
            else:
                raise Exception("当前系统不支持自动转换，请手动转换")

    @staticmethod
    def deal_err_xlsx(xlsx_file):
        """
        如果有无法正常打开的xlsx文件，通过此程序转换后打开。(ARGB报错问题)
        todo 操作完成后，删除 tmp 文件，避免二次执行的异常
        todo 遇到保存 xls 后 xls2xlsx 转换失败的问题
        1. 本来是 xlsx 文件，但是重命名成 xls 文件
        2. 本来是 xls 文件，但是重命名成 xlsx 文件
        """
        from RPA.Excel.Files import Files
        cp_xlsx_file = os.path.splitext(os.path.basename(xlsx_file))[0]+ '_cp.xlsx'
        tmp_xlsx = os.path.splitext(os.path.basename(xlsx_file))[0]+ '_tmp.xlsx'
        tmp_xls = os.path.splitext(os.path.basename(xlsx_file))[0] + '_tmp.xls'
        shutil.copy2(xlsx_file, cp_xlsx_file) # 复制一份数据

        try:
            app = Files()
            os.rename(xlsx_file, tmp_xlsx)
            app.open_workbook(tmp_xlsx)
            app.save_workbook(tmp_xls)  # 使用rpaframework打开，并保存为xls文件（保存xlsx文件后，openpyxl无法打开）
            time.sleep(3)
            excel.xls2xlsx(tmp_xls, xlsx_file)  # 将xls文件转换成xlsx文件，会丢失格式、列宽等
            time.sleep(2)
        except:
            try:
                tmp_xlsx = os.path.splitext(os.path.basename(xlsx_file))[0] + '_tmp.xlsx'
                tmp_xls = os.path.splitext(os.path.basename(xlsx_file))[0] + '_tmp.xls'
                app = Files()
                app.open_workbook(tmp_xlsx)
                app.save_workbook(xlsx_file)  # 使用rpaframework打开，并保存为xls文件（保存xlsx文件后，openpyxl无法打开）
                time.sleep(3)
                time.sleep(2)
                app.close_workbook()
            except:
                pass
                print("")
        finally:
            if os.path.exists(tmp_xlsx):
                os.remove(tmp_xlsx)
            if os.path.exists(tmp_xls):
                os.remove(tmp_xls)
            if not os.path.exists(xlsx_file):
                os.rename(cp_xlsx_file, xlsx_file)

    @staticmethod
    def activate_formula(file_name):
        """
        打开并保存excel文件，用于激活excel文件中的公式，或者处理一下非标的xlsx文件
        使用后边框属性会变化
        """
        sysinfo = FRAME.get_operating_system()
        exec_env = FRAME.get_env_arg("RPA_ENV", "")
        if exec_env == "prod" or exec_env == 'test':  # 服务器环境执行
            result = subprocess.run(
                ["/opt/libreoffice25.2/program/python", "CLPC/uno_script.py", "opensave", file_name],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print("激活公式过程日志：", result.stdout.decode())
            if result.returncode != 0:
                print("激活公式过程异常：", result.stderr.decode())

                raise Exception("LibreOffice激活公式失败")
            else:
                print("LibreOffice激活公式成功")
        else:
            if sysinfo == "MacOS":
                result = subprocess.run(
                    ["/Applications/LibreOffice.app/Contents/Resources/python", "CLPC/uno_script.py", "opensave",
                     file_name],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                print("激活公式过程日志：", result.stdout.decode())
                if result.returncode != 0:
                    print("激活公式过程异常：", result.stderr.decode())

                    raise Exception("LibreOffice激活公式失败")
                else:
                    print("LibreOffice激活公式成功")

            elif sysinfo == "Linux":
                pass
                # 信创环境后续适配
            elif sysinfo == "Windows":
                from win32com.client import Dispatch
                try:
                    xlApp = Dispatch("Ket.Application")  # 使用wps， 如果使用office，需要改为 "Excel.Application"
                except:
                    raise Exception("请确认是否安装WPS软件")
                try:
                    xlApp.Visible = False
                    xlApp.DisplayAlerts = False
                    dir_name = os.path.dirname(file_name)
                    if "\\" in file_name or "/" in file_name or dir_name != "":
                        pass
                    else:
                        main_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
                        file_name = os.path.join(main_dir, file_name)
                    xlApp.Workbooks.Open(file_name)
                    xlApp.ActiveWorkbook.Save()
                    xlApp.ActiveWorkbook.Close()
                    xlApp.Quit()
                except:
                    raise Exception("请确认文件存在，或者文件没有被其他程序占用")

            else:
                raise Exception("当前系统不支持自动转换，请手动转换")

    @classmethod
    def open_xlsx_through_office(cls, file_name, data_only=False):
        """
         todo 通过libreoffice或者wps打开excel文件，然后另存为xlsx格式，保证文件可以被正确处理
        :param file_name: 文件路径
        :param data_only: 是否只读取数据
        """
        sysinfo = FRAME.get_operating_system()
        exec_env = FRAME.get_env_arg("RPA_ENV", "")
        if exec_env == "prod" or exec_env == 'test':  # 服务器环境执行
            result = subprocess.run(
                ["/opt/libreoffice25.2/program/python", "CLPC/uno_script.py", "opensave", file_name],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print("另存为过程日志：", result.stdout.decode())
            if result.returncode != 0:
                print("另存为过程异常：", result.stderr.decode())

                raise Exception("LibreOffice另存为失败")
            else:
                print("LibreOffice另存为成功")
        else:
            if sysinfo == "MacOS":
                os.system(
                    "/Applications/LibreOffice.app/Contents/Resources/python CLPC/uno_script.py opensave {}".format(file_name))
            elif sysinfo == "Linux":
                pass
                # 信创环境后续适配
            elif sysinfo == "Windows":
                from win32com.client import Dispatch
                try:
                    xlApp = Dispatch("Ket.Application")  # 使用wps， 如果使用office，需要改为 "Excel.Application"
                    xlApp.Visible = False
                    xlApp.DisplayAlerts = False
                    abs_name = os.path.abspath(file_name)
                    xlApp.Workbooks.Open(abs_name)
                    xlApp.ActiveWorkbook.Save()
                    xlApp.ActiveWorkbook.Close()
                    xlApp.Quit()
                except:
                    print("请确认是否安装WPS软件")
            else:
                raise Exception("当前系统不支持自动转换，请手动转换")
        try:
            workbook = load_workbook(abs_name, data_only=data_only)  # 默认读取公式而非值
        except:
            raise Exception("文件打开失败，请使用office软件创建标准xlsx文件作为模板")
        # 返回excel的实例
        self = cls.__new__(cls)  # 使用__new__创建实例
        self.__dict__.update(workbook.__dict__)  # 复制workbook的属性到self
        return self

    @staticmethod
    def open_excel(file_name, data_only=False):
        name, excel_type = os.path.splitext(os.path.basename(file_name))
        if excel_type == '.xls':
            if excel.check_excel_file_format(file_name):
                print("打开xls格式的表格，将通过兼容模式处理")
                xlsx_name = name + '.xlsx'
                excel.xls2xlsx(file_name, xlsx_name)
                return excel.open_xlsx(xlsx_name, data_only=data_only)
            else:
                new_file_name = name + ".xlsx"
                os.rename(file_name, new_file_name)
                time.sleep(2)
                try:
                    return excel.open_xlsx(new_file_name, data_only=data_only)
                except:
                    print("使用兼容模式打开此文件，格式可能有变化")
                    return excel.open_xlsx(new_file_name, compatibility_mode=True, data_only=data_only)

        elif excel_type == '.xlsx':
            if excel.check_excel_file_format(file_name):
                try:
                    return excel.open_xlsx(file_name, data_only=data_only)
                except:
                    print("使用兼容模式打开此文件，格式可能有变化")
                    return excel.open_xlsx(file_name, compatibility_mode=True, data_only=data_only)
            else:
                new_file_name = name + ".xls"
                os.rename(file_name, new_file_name)
                print("打开xls格式的表格，将通过兼容模式处理")
                xlsx_name = name + '.xlsx'
                excel.xls2xlsx(file_name, xlsx_name)
                return excel.open_xlsx(xlsx_name, data_only=data_only)
        else:
            print("当前待打开文件名：", file_name)
            raise Exception('当前文件格式不支持，请使用xls，xlsx格式的表格文件')

    @staticmethod
    def open_excel_with_data(file_name):
        """
        以获取公式结果的方式，打开excel文件
        """
        name, excel_type = os.path.splitext(os.path.basename(file_name))
        if excel_type == '.xls':
            if excel.check_excel_file_format(file_name):
                print("打开xls格式的表格，将通过兼容模式处理")
                xlsx_file_name = excel.convert_xls_file(file_name)
                excel.activate_formula(xlsx_file_name)
                return excel.open_xlsx(xlsx_file_name, data_only=True)
            else:
                print("请确认excel文件后缀是否准确")

        elif excel_type == '.xlsx':
            excel.activate_formula(file_name)
            sleep(2)
            return excel.open_excel(file_name, data_only=True)
        else:
            print("当前待打开文件名：", file_name)
            raise Exception('当前文件格式不支持，请使用xls，xlsx格式的表格文件')

    @classmethod
    def open_patn_excel(cls, file_name):
        """
        用于打开模板文件，直接使用openpyxl打开，
        """
        try:
            workbook = load_workbook(file_name, data_only=False)  # 默认读取公式而非值
        except:
            raise Exception("文件打开失败，请使用office软件创建标准xlsx文件作为模板")
        # 返回excel的实例
        self = cls.__new__(cls)  # 使用__new__创建实例
        self.__dict__.update(workbook.__dict__)  # 复制workbook的属性到self
        return self

    @classmethod
    def decrypt_xlsx(cls, file_name, pwd):
        """
        解密已加密的文件，另存为一份不加密的文件，并返回不加密的文件名
        """
        sysinfo = FRAME.get_operating_system()
        exec_env = FRAME.get_env_arg("RPA_ENV", "")

        name, excel_type = os.path.splitext(os.path.abspath(file_name))
        output_file = name + '_decrypt' + excel_type

        if exec_env == "prod" or exec_env == 'test':  # 服务器环境执行
            result = subprocess.run(
                ["/opt/libreoffice25.2/program/python", "CLPC/uno_script.py", "decrypt", file_name, pwd],
                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print("去除密码过程日志：", result.stdout.decode())
            if result.returncode != 0:
                print("去除密码过程异常：", result.stderr.decode())

                raise Exception("LibreOffice去除密码失败")
            else:
                print("LibreOffice去除密码成功")
        else:
            if sysinfo == "MacOS":
                convert_result = subprocess.run(
                    ["/Applications/LibreOffice.app/Contents/Resources/python", "CLPC/uno_script.py", "decrypt",
                     file_name, pwd],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if convert_result.returncode != 0:
                    print("去除密码过程日志：", convert_result.stdout.decode())
                    raise Exception("LibreOffice去除密码失败: {}".format(convert_result.stderr.decode()))
                else:
                    print("LibreOffice去除密码成功")
            elif sysinfo == "Linux":
                pass
                # 信创环境后续适配
            elif sysinfo == "Windows":
                from win32com.client import Dispatch
                import msoffcrypto
                from msoffcrypto import OfficeFile
                try:

                        # 加载加密的Excel文件
                    office_file = OfficeFile(file_name)
                        # 设置密码
                    office_file.load_key(password=pwd)
                        # 解密并保存到新文件
                    office_file.decrypt_write(output_file)


                except:
                    print("请确认是否安装WPS软件")
            else:
                raise Exception("当前系统不支持自动转换，请手动转换")

        return output_file

    @classmethod
    def create_excel(cls):
        """
        新建excel文档，并返回excel实例。
        默认包含一个sheet页，名称为Sheet
        """
        workbook = Workbook()
        self = cls.__new__(cls)  # 使用__new__创建实例
        self.__dict__.update(workbook.__dict__)  # 复制workbook的属性到self
        return self


    #打开文档
    @classmethod
    # @class_func_tracker
    def open_xlsx(cls,filename, compatibility_mode=False, data_only=False, read_only=False):
        """
        打开一个已存在的Excel文件，并返回excel实例。
        """
        # 使用openpyxl的load_workbook方法加载工作簿
        if compatibility_mode:
            try:
                excel.deal_err_xlsx(filename)
            except:
                raise Exception("当前文件可能有损坏，暂不支持自动处理，可以手动另存为后再进行处理")
        time.sleep(1)
        try:
            workbook = load_workbook(filename, data_only=data_only, read_only=read_only)
        except:
            raise Exception("文件打开失败，请尝试使用兼容模式打开")
        # 返回excel的实例
        self = cls.__new__(cls)  # 使用__new__创建实例
        self.__dict__.update(workbook.__dict__)  # 复制workbook的属性到self
        return self

    #获取sheet
    def get_sheet(self, sheet_name):
        """获取指定名称的工作表，并复制其属性到self"""
        sheet = self[sheet_name]
        print(sheet)
        sht = MyWorksheet(sheet)
        return MyWorksheet(sheet)

    def get_current_sheet(self):
        """
        获取默认激活的sheet表
        """
        sht = self.active
        return MyWorksheet(sht)

    #创建新sheet
    def add_sheet(self,sheet_name) -> Worksheet:
        sheet = self.create_sheet(sheet_name)
        return sheet

    #删除指定sheet
    def remove_sheet(self,sheet_name):
        sheet=self.remove(self[sheet_name])
        return sheet

    #获取excel中所有sheet
    def get_sheetnames(self):
        sheet_names=self.sheetnames
        return sheet_names
    @staticmethod
    def copy_cell_with_style(source_cell: Cell, target_cell: Cell):
        """
        带格式复制单元格
        :param source_cell 原始单元格
        :param target_cell 目标单元格
        """
        # if source_cell.has_style:
        target_cell.value = copy.copy(source_cell.value)
        target_cell.style = copy.copy(source_cell.style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)

    @staticmethod
    def style_brush(source_cell: Cell, target_cell: Cell):
        """
        格式刷
        :param source_cell 原始单元格
        :param target_cell 目标单元格
        """
        target_cell.style = copy.copy(source_cell.style)
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)

    @staticmethod
    def copy_area_with_style(source_sht, source_range: str, target_sht, target_start_cell:str):
        """
        依次复制区域内所有单元格
        :param source_sht: 源sheet（被拷贝的sheet）
        :param source_range: 源区域（被拷贝的区域）
        :param target_sht: 目标sheet（拷贝到的sheet）
        :param target_start_cell: 目标起始单元格
        """
        start_col, start_row, end_col, end_row = range_boundaries(source_range)
        target_start_row, target_start_col = coordinate_to_tuple(target_start_cell)

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                source_cell = source_sht.sheet.cell(row=row, column=col)
                target_cell = target_sht.sheet.cell(row=target_start_row + row - start_row, column=target_start_col + col - start_col)
                excel.copy_cell_with_style(source_cell, target_cell)

    @staticmethod
    def join_sheets_vertical(final_sht, joined_sht):
        """
        合并两个sheet，将joined_sht的内容合并到final_sht中
        竖向合并
        默认拼接到 final_sht 的最后一行
        :param final_sht: 最终合并后的sheet
        :param joined_sht: 需要被合并的sheet
        """
        joined_range = joined_sht.get_all_data_range()
        final_last_row = final_sht.row_count()
        target_cell = "A" + str(final_last_row + 1)
        excel.copy_area_with_style(joined_sht, joined_range, final_sht, target_cell)

    @staticmethod
    def join_sheets_horizontal(final_sht, joined_sht):
        """
        合并两个sheet，将joined_sht的内容合并到final_sht中
        横向合并
        默认拼接到 final_sht 的最后一列
        :param final_sht: 最终合并后的sheet
        :param joined_sht: 需要被合并的sheet
        """
        joined_range = joined_sht.get_all_data_range()
        final_last_col = final_sht.col_count()
        target_cell = get_column_letter(final_last_col + 1) + "1"
        excel.copy_area_with_style(joined_sht, joined_range, final_sht, target_cell)

    def __del__(self):
        self.close()  # 退出时关闭excel

    @staticmethod
    def convert_excel_to_img(excel_file_path:str, sheet_name:str, png_file_path:str, timeout=60):
        """
        通过服务将excel文件转化成图片
        :param excel_file_path: excel文件路径
        :param sheet_name: sheet名称
        :param png_file_path: 保存的图片路径 后缀需要为png
        :param timeout: 请求超时时间，默认60秒
        """
        import requests
        import base64
        from io import BytesIO
        from PIL import Image
        import re
        excel.activate_formula(excel_file_path)  # 激活公式


        api_url = FRAME.BACKEND_ADDR_PROD + "api/base/excelToImage"  # 因测试服务器分公司无法访问，因此暂时统一使用生产地址

        data = {'sheetName': sheet_name}

        # 打开文件
        with open(excel_file_path, 'rb') as file:
            files = {'file': file}

            # 发送POST请求
            try:
                response = requests.post(api_url, files=files, data=data, timeout=timeout)
            except:
                print("请求超时, 请稍后再试")
                raise Exception("请求超时")

        # 检查响应状态码
        if response.status_code == 200:
            res_data = response.json()
            res_status = res_data.get("code")
            if res_status == 1:
                png_b64 = res_data.get("data")

                result = re.search("data:image/(?P<ext>.*?);base64,(?P<data>.*)", png_b64, re.DOTALL)
                if result:
                    data = result.groupdict().get("data")

                img_data = base64.b64decode(data)
                # 将字节数据转换为图片并保存
                image = Image.open(BytesIO(img_data))
                image.save(png_file_path)

                print("图片生成成功")
            else:
                res_msg = res_data.get("msg")
                print(f"图片生成失败，原因: {res_msg}")
        else:
            print(f"文件上传失败，状态码: {response.status_code}")

    @staticmethod
    def convert_xlsx_2_png(xlsx_file: str, sheet_name: str, png_file_path: str):
        """
        使用 libreoffice 将 xlsx 文件转换为 png 图片
        :param xlsx_file: 待截图的xlsx文件路径
        :param sheet_name: 待截图sheet页名
        :param png_file_path: 生成截图文件名，需为png后缀名
        :return: 生成的png图片名
        """

        # 2. 定义 size 到像素的映射（根据常见标准）
        SIZE_TO_PX = {
            1: "10px",
            2: "13px",
            3: "16px",  # LibreOffice 默认 size="3" 约 12pt → 16px
            4: "18px",
            5: "24px",
            6: "32px",
            7: "48px"
        }

        def calculate_total_height(soup):
            """
            """
            total_height = 0

            td_tags = soup.find_all('td', height=True)
            for td in td_tags:
                try:
                    total_height += int(td['height'])
                except ValueError:
                    continue

            return total_height

        def calculate_total_width(soup):
            """
            """
            total_width = 0

            colgroup_tags = soup.find_all('colgroup', width=True)
            for colgroup in colgroup_tags:
                try:
                    try:
                        span = colgroup['span']
                        total_width += int(colgroup['width']) * int(span)
                    except:
                        total_width += int(colgroup['width'])
                except ValueError:
                    continue

            return total_width

        def convert_font_to_style(soup):
            #查找并转换所有 < font > 标签
            for font_tag in soup.find_all('font'):
                style_parts = []  # 存储 CSS 样式片段

                # 处理 face 属性 → font-family
                if 'face' in font_tag.attrs:
                    font_family = font_tag['face']
                    # 添加引号确保带空格的字体名有效
                    style_parts.append(f"font-family: '{font_family}'")
                    del font_tag['face']  # 删除旧属性

                # 处理 size 属性 → font-size
                if 'size' in font_tag.attrs:
                    size_val = font_tag['size']
                    try:
                        # 处理相对值（如 size="+1"）
                        if size_val.startswith(('+', '-')):
                            base_size = 3  # 默认基准 size=3 (16px)
                            adjusted_size = base_size + int(size_val)
                            px_size = SIZE_TO_PX.get(adjusted_size, "16px")
                        else:
                            px_size = SIZE_TO_PX.get(int(size_val), "16px")
                        style_parts.append(f"font-size: {px_size}")
                    except ValueError:
                        pass  # 忽略无效值
                    del font_tag['size']  # 删除旧属性
                else: # 没有size属性，默认给出16px
                    style_parts.append(f"font-size: 16px")

                # 处理 color 属性 → color
                if 'color' in font_tag.attrs:
                    style_parts.append(f"color: {font_tag['color']}")
                    del font_tag['color']  # 删除旧属性

                # 4. 替换为 <span> 并添加样式
                font_tag.name = 'span'  # 修改标签名为 span
                if style_parts:
                    # 合并样式并保留原有 style（如有）
                    existing_style = font_tag.get('style', '')
                    new_style = existing_style + '; ' if existing_style else ''
                    new_style += '; '.join(style_parts)
                    font_tag['style'] = new_style
            return soup

        from bs4 import BeautifulSoup
        sysinfo = FRAME.get_operating_system()
        exec_env = FRAME.get_env_arg("RPA_ENV", "")
        if exec_env == "prod" or exec_env == 'test': # 服务器环境执行
            convert_result = subprocess.run(["/opt/libreoffice25.2/program/python", "CLPC/uno_script.py", "convert_html", xlsx_file],
                                            stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            if convert_result.returncode != 0:
                print("文件转换过程日志：", convert_result.stdout.decode())
                raise Exception("LibreOffice转换失败: {}".format(convert_result.stderr.decode()))
            else:
                print("LibreOffice转换成功")

            html_file = xlsx_file.replace(".xlsx", ".html")
            sht_html = sheet_name + '.html'

            # 根据给定的sheet名称，提取html中对应的table
            with open(html_file, 'r') as file:  # 读取转换后的html文件
                html_content = file.read()

                content_soup = BeautifulSoup(html_content, "html.parser")
                try:
                    em_content = content_soup.find("em", text=sheet_name)
                except:
                    raise Exception(f"sheet页：{sheet_name}不存在，请确认正确的名称")
                try:
                    a_content = em_content.find_parents("a")[0]
                    sht_content = a_content.find_next_sibling("table")
                except:
                    sht_content = content_soup.find("table")

                sht_content = convert_font_to_style(sht_content)
                with open(sht_html, 'w', encoding='utf-8') as sht_file:  # 写入只有所需sheet的html
                    sht_file.write(str(sht_content))

            TOOL.html_table_2_png(sht_html, png_file_path)

        else:
            if sysinfo == "MacOS" :

                convert_result = subprocess.run(["/Applications/LibreOffice.app/Contents/Resources/python", "CLPC/uno_script.py", "convert_html", xlsx_file, sheet_name],
                                                stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if convert_result.returncode != 0:
                    print("文件转换过程日志：", convert_result.stdout.decode())
                    raise Exception("LibreOffice转换失败: {}".format(convert_result.stderr.decode()))
                else:
                    print("LibreOffice转换成功")

                html_file = xlsx_file.replace(".xlsx", ".html")
                sht_html = sheet_name + '.html'

                # 根据给定的sheet名称，提取html中对应的table
                with open(html_file, 'r') as file: # 读取转换后的html文件
                    html_content = file.read()

                    content_soup = BeautifulSoup(html_content, "html.parser")
                    try:
                        em_content = content_soup.find("em", text=sheet_name)
                    except:
                        raise Exception(f"sheet页：{sheet_name}不存在，请确认正确的名称")
                    try:
                        a_content = em_content.find_parents("a")[0]
                        sht_content = a_content.find_next_sibling("table")
                    except:
                        sht_content = content_soup.find("table")

                    sht_content = convert_font_to_style(sht_content)
                    with open(sht_html, 'w', encoding='utf-8') as sht_file:  # 写入只有所需sheet的html
                        sht_file.write(str(sht_content))

                TOOL.html_table_2_png(sht_html, png_file_path)

            elif sysinfo == "Linux":
                pass
                # 信创环境后续适配
            elif sysinfo == "Windows":
                #
                excel.convert_excel_to_img(xlsx_file, sheet_name, png_file_path)
                pass
            else:
                raise Exception("当前系统不支持自动转换，请手动转换")

@FUNC_USAGE_TRACKER
class MyWorksheet:
    def __init__(self, sheet:Worksheet):
        # 复制传入的Worksheet对象到self
        self.sheet=sheet

    #读取内容到单元格
    def read(self, area=None):
        """
        读取当前激活的sheet的数据，支持读取单元格、行、列和区域
        """
        try:
            if area==None:
                # 初始化一个空列表来存储所有单元格的数据
                all_cells_data = []
                # 遍历工作表中的所有行

                for row in self.sheet.iter_rows(min_row=1, max_row=self.sheet.max_row, min_col=1, max_col=self.sheet.max_column):
                    # 对于每一行，获取所有单元格的值，并添加到行列表中
                    row_data = [cell.value for cell in row]
                    all_cells_data.append(row_data)
                return all_cells_data
            elif area.isalpha():
                col = self.sheet[area]
                col_data = []
                for cell in col:
                    col_data.append(cell.value)
                #把一维数组转化成二维数组
                cols = [[item] for item in col_data]
                return cols

            elif area.isdigit():
                row = self.sheet[area]
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                return row_data

            elif ':' in area:
                # 使用切片语法获取指定范围的所有单元格
                cells = self.sheet[area]
                # 使用列表推导式遍历单元格范围，并将值收集到一个列表中
                # 这里使用嵌套的列表推导式来创建一个二维列表
                data = [[cell.value for cell in row] for row in cells]
                return data

            elif any(c.isalpha() for c in area) and any(c.isdigit() for c in area):

                # 尝试通过名称获取工作表
                return self.sheet[area].value

            else:
                print('输入参数格式有误')

        except KeyError:
            return None

    def read_all_data(self, start_row=1) -> list:
        """
        读取当前sheet的全量数据，返回数组
        :param start: 数据起始行，默认第一行
        """
        range = self.get_all_data_range(start_row=start_row)
        result = self.read(area=range)
        return result

    #写入内容到单元格
    def write(self,start_cell, values: Union[List[List[Any]], List[Any], str]):
        """
        将一维或二维数组写入Excel工作表的指定起始单元格。

        :param sheet: openpyxl的工作表对象
        :param start_cell: 写入的起始单元格位置，如 'B1'
        :param values: 数组，可以是一维或二维，包含要写入的数据
        """
        # 检查value是否为列表，如果不是，转换为单元素列表
        if not isinstance(values, list):
            values = [values]
        # 如果values是一维数组，转换为二维数组（单行）
        if not all(isinstance(row, list) for row in values):
            values = [values]

        # 获取起始单元格的列字母和行号
        # start_col_letter = start_cell[0]
        # start_row = int(start_cell[1:])
        start_row, start_col = coordinate_to_tuple(start_cell)

        # 初始化列索引
        col_idx = 0
        # 遍历values中的每个元素
        for row in values:
            for value in row:
                # 计算目标单元格的列字母
                col_letter = get_column_letter(start_col + col_idx)
                # 写入数据到指定单元格
                self.sheet[col_letter + str(start_row)] = value
                # 更新列索引
                col_idx += 1
            # 重置列索引，因为新的行将从第一列开始
            col_idx = 0
            # 更新行号，因为要写入下一行
            start_row += 1

    def get_all_data_range(self, start_row=1) -> str:
        """
        获取当前sheet所有数据的范围
        :param start_row:起始行，默认1为第一行
        """
        row = self.row_count()
        col = self.col_count()
        col_alpha = get_column_letter(col)
        result = "A{}:{}{}".format(start_row, col_alpha, row)
        return result
        pass

    #获取行数
    def row_count(self):
        cells = self.sheet._cells
        max_row = 1
        data_cell = {}
        for (c,v) in cells.items():
            if v.value != None:
                data_cell.update({c:v})
        if data_cell:
            max_row =  max(data_cell)[0]
        return max_row

    #获取列数
    def col_count(self):
        max_col = 1
        cells = self.sheet._cells
        data_cell = {}
        for (c, v) in cells.items():
            if v.value != None:
                data_cell.update({c: v})
        if data_cell:
            max_col = max(c[1] for c in data_cell)
        return max_col


    #合并单元格
    def merge_cell(self,range):
        self.sheet.merge_cells(range)

    #拆分已合并单元格
    def unmerge_cell(self,range):
        self.sheet.unmerge_cells(range)

    def unmerge_all_merged_cells(self):
        """
        拆分所有合并单元格
        """
        merged_ranges = copy.deepcopy(self.sheet.merged_cells.ranges)
        for merged_cell in merged_ranges:
            self.sheet.unmerge_cells(merged_cell.coord)


    def unmerge_and_fill(self):
        '''
        拆分并填充所有合并单元格
        '''
        sheet=self.sheet
        # 复制合并单元格范围，避免在遍历过程中修改原范围导致错误
        merged_ranges = copy.deepcopy(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            # 获取合并单元格的边界信息
            min_col, min_row, max_col, max_row = merged_range.bounds
            # 获取合并单元格左上角的值
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            # 拆分合并单元格
            sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            # 遍历拆分后的每个单元格并填充值
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col, value=top_left_value)

    def find(self, keyword, mode:str="equal") -> list:
        """
        查找内容所在位置，返回位置列表
        :param keyword: 查找关键字
        :param mode: 查找模式，支持相等、包含；equal，contain
        :return : 返回所有匹配单元的坐标列表，形式：[(1, "A"),(2,"C")]，没有匹配返回空列表
        """
        positions = []
        for row in self.sheet.iter_rows():
            for cell in row:
                if mode == "equal":
                    if cell.value == keyword:
                        p = f'{cell.column_letter}{cell.row}'
                        positions.append((p, str(cell.row), cell.column_letter))
                if mode == 'contain':
                    if str(keyword) in str(cell.value) and cell.value:
                        p = f'{cell.column_letter}{cell.row}'
                        positions.append((p, str(cell.row), cell.column_letter))
        return positions

    def find_first(self, keyword, mode:str="equal") -> (int,str):
        """
        查找内容所在第一个位置，返回所在行列,查找顺序，从上到下，做左到右
        :param keyword: 查找关键字
        :param mode: 查找模式，支持相等、包含；equal，contain
        :return: row,col 第一个匹配的单元格的行号和列号，未找到则返回None,None。
        """
        for row in self.sheet.iter_rows():
            for cell in row:
                if mode == "equal":
                    if cell.value == keyword and cell.value:
                        result_row = cell.row
                        result_col = cell.column_letter
                        p = f'{cell.column_letter}{cell.row}'
                        return p, str(result_row), result_col
                if mode == 'contain':
                    if str(keyword) in str(cell.value) and cell.value:
                        result_row = cell.row
                        result_col = cell.column_letter
                        p = f'{cell.column_letter}{cell.row}'
                        return p, str(result_row), result_col
        else:
            print(f"未找到关键字:{keyword}")
            return None, None


    #插入行或列
    def insert(self,row_or_col):
        range=str(row_or_col)
        if range.isalpha():
            column_number = column_index_from_string(range)
            self.sheet.insert_cols(column_number)
        elif range.isdigit():
            row_number = int(range)
            self.sheet.insert_rows(row_number)
        else:
            print('输入参数格式错误')

    #删除行或列
    def delete(self,row_or_col):
        range=str(row_or_col)
        if range.isalpha():
            column_number = column_index_from_string(range)
            self.sheet.delete_cols(column_number)
        elif range.isdigit():
            row_number = int(range)
            self.sheet.delete_rows(row_number)
        else:
            print('输入参数格式错误')

    def delete_rows(self, start_row:int, end_row:int):
        """
        批量删除行
        :param start_row: 起始行号
        :param end_row: 结束行号
        """
        row_cnt = end_row - start_row + 1
        self.sheet.delete_rows(start_row, row_cnt)

    def delete_cols(self, start_col:str, end_col:str):
        """
        批量删除列
        :param start_col:起始列号
        :param end_col:结束列号
        """
        start_col_num = column_index_from_string(start_col)
        end_col_num = column_index_from_string(end_col)
        col_cnt = end_col_num - start_col_num + 1
        self.sheet.delete_cols(start_col_num, col_cnt)

            # 定义一个函数来复制样式

    def get_cell(self, c: str) -> Cell:
        """
        获取单元格对象，包含内容与格式信息
        :param c: 单元格坐标
        """
        result_cell = self.sheet[c]
        return result_cell

    def sort_data_by_col(self, data, col: str, sort_ascending: bool = True) -> Union[List[any], List[List[any]]]:
        """
        根据固定列进行排序
        :param data: 待排序的数据，使用read进行读取
        :param col: 排序列号
        :param sort_ascending: 升序为True，降序为False
        """
        from RPA.Tables import Tables
        tb = Tables()
        tb_data = tb.create_table(data)  # 从二维数组转可用TB
        col_no = column_index_from_string(col) - 1  # 从0开始
        tb.sort_table_by_column(tb_data, col_no, sort_ascending)  # 排序会带首行，需要自己来控制
        sort_result_transposed = list(tb.export_table(tb_data, with_index=False, as_list=False).values())
        result = [list(i) for i in zip(*sort_result_transposed)]
        return result

    def sort_data_by_col_by_sheet(self, start_row: int, col: str, sort_ascending: bool = True) -> Union[List[any], List[List[any]]]:
        """
        根据固定列进行排序，无需传入data，自动读取
        :param start_row: 输入起始行
        :param col: 排序列号
        :param sort_ascending: 升序为True，降序为False
        """
        from RPA.Tables import Tables
        data_range = self.get_all_data_range(start_row)
        data = self.read(area=data_range)
        tb = Tables()
        tb_data = tb.create_table(data)  # 从二维数组转可用TB
        col_no = column_index_from_string(col) - 1  # 从0开始
        tb.sort_table_by_column(tb_data, col_no, sort_ascending)  # 排序会带首行，需要自己来控制
        sort_result_transposed = list(tb.export_table(tb_data, with_index=False, as_list=False).values())
        result = [list(i) for i in zip(*sort_result_transposed)]
        return result

    def filter_data_by_col(self, data, col:str, filter_operator:str, filter_value:str) -> Union[List[any], List[List[any]]]:
        """
        根据固定列进行筛选
        :param data: 待筛选的数据，使用read进行读取
        :param col: 筛选列号
        :param filter_operator: >, <, >=, <=, ==, !=, contains
        :param filter_value:
        """
        from RPA.Tables import Tables
        tb = Tables()
        tb_data = tb.create_table(data)  # 从二维数组转可用TB
        col_no = column_index_from_string(col) - 1  # 从0开始

        if filter_operator != 'contains' and filter_operator != 'not contains' and filter_operator != "!=" and filter_operator != "==":
            try:
                filter_value = float(filter_value)
            except:
                pass
        tb.filter_table_by_column(tb_data, col_no, filter_operator, filter_value)  # todo >,< 等 value应该要处理为数字  Tables源码需要更新，并打包到基础环境中

        filter_result_transposed = list(tb.export_table(tb_data, with_index=False, as_list=False).values())
        result = [list(i) for i in zip(*filter_result_transposed)]
        return result

    def filter_data_by_col_by_sheet(self, start_row:int, col:str, filter_operator:str, filter_value:str) -> Union[List[any], List[List[any]]]:
        """
        根据固定列进行筛选
        :param start_row: 输入起始行
        :param col: 筛选列号
        :param filter_operator: >, <, >=, <=, ==, !=, contains
        :param filter_value: 筛选值
        """
        from RPA.Tables import Tables
        tb = Tables()
        data = self.get_all_data_range(start_row)
        tb_data = tb.create_table(data)  # 从二维数组转可用TB
        col_no = column_index_from_string(col) - 1  # 从0开始

        if filter_operator != 'contains' and filter_operator != 'not contains':
            try:
                filter_value = float(filter_value)
            except:
                pass
        tb.filter_table_by_column(tb_data, col_no, filter_operator, filter_value)  # todo >,< 等 value应该要处理为数字  Tables源码需要更新，并打包到基础环境中

        filter_result_transposed = list(tb.export_table(tb_data, with_index=False, as_list=False).values())
        result = [list(i) for i in zip(*filter_result_transposed)]
        return result

    def get_col_unique_value(self, col:str, exclude_value:list =[] ) -> List:
        """
        对指定列的数据进行去重，并返回去重后的数据
        :param col 列序号
        :param exclude_value 需要排除的值，如列标题
        :return 返回去重后的结果列
        """
        col_value = self.read(col)  # 读取列的时候，是个二维数组
        col_value_1d = sum(col_value, [])
        unique_values_set = set(col_value_1d)
        if len(exclude_value) > 0:
            for ev in exclude_value:
                unique_values_set.discard(ev)
        unique_values = list(unique_values_set)
        return unique_values

    def split_data_by_value(self, start_row:int, value, col:str, split_mode='equal'):
        """
        根据指定列的特征值拆分数据，返回拆分后的数据
        :param start_row
        :param value
        :param col
        :param split_mode 筛选特征值的方式：equal、contain
        :return 拆分后数据
        """
        data = self.read_all_data(start_row)
        col_no = column_index_from_string(col) - 1  # 从0开始
        result_data = []
        for row_info in data:
            cell_value = row_info[col_no]
            if split_mode == 'equal':
                if cell_value == value:
                    result_data.append(row_info)
            elif split_mode == 'contain':
                if value in cell_value:
                    result_data.append(row_info)
            else:
                raise Exception("拆分模式错误，请确认入参")
        return result_data

    def set_cell_numeric(self, cell:Cell):
        """
        设置单元格为数字格式
        """
        from openpyxl.styles import numbers
        if cell.data_type == 'n':
            return

        val = cell.value
        val_num_str = val.replace(',', '').replace('%', '').replace(' ', '')
        try:
            val_num = float(val_num_str)
        except:
            print("不是数字格式")
            return

        val_num_str = val.replace(',', '').replace(' ', '')
        if '%' in val:
            val_num = float(val_num_str.replace('%', ''))/100.0

        cell.data_type = 'n'
        cell.number_format = numbers.FORMAT_GENERAL
        cell.value = val_num
        cell.data_type = 'n'


    def set_col_numeric(self, col:str):
        """
        设置列为数字格式
        """
        col_cells = self.sheet[col]
        for cell in col_cells:
            self.set_cell_numeric(cell)

    def calculate_col(self, col:str, calculate_type:str, start_row=1):
        """
        对列进行简单计算，表格需从第一行开始填，部分表格有空行的，不适用于此方法
        适用于求和、平均值、最大值、最小值，将在对应列的最后一行写入计算结果
        :param col: 列号
        :param calculate_type: 总计, 平均值, 最大值, 最小值
        :param start_row: 起始行
        """
        data_col = self.sheet[col]
        data_cell = []
        max_row = 1
        for cell in data_col:
            if cell.value != None:
                data_cell.append(cell.row)
        if data_cell:
            max_row = max(data_cell)  # 计算实际数据共多少行

        end_row = max_row + 1
        target_coord = col + str(end_row)
        formula_end_coord = col + str(max_row)

        if calculate_type == '总计':
            formula = '=SUM({}:{})'.format(col + str(start_row), formula_end_coord)
        elif calculate_type == '平均值':
            formula = '=AVERAGE({}:{})'.format(col + str(start_row), formula_end_coord)
        elif calculate_type == '最大值':
            formula = '=MAX({}:{})'.format(col + str(start_row), formula_end_coord)
        elif calculate_type == '最小值':
            formula = '=MIN({}:{})'.format(col + str(start_row), formula_end_coord)
        else:
            raise Exception("计算类型错误")
        self.write(target_coord, formula)

    def calculate_sht(self, calculate_type:str, start_row=1, start_col="A"):
        """
        对整个sheet进行简单计算，表格需从第一行开始填，部分表格有空行的，不适用于此方法 \n
        适用于求和、平均值、最大值、最小值，将在最后一行写入计算结果
        默认会计算到最后一列
        :param calculate_type: 总计, 平均值, 最大值, 最小值
        :param start_row: 数据起始行
        :param start_col: 数据起始列
        """
        start_col_num = column_index_from_string(start_col)
        for col in range(start_col_num, self.col_count() + 1):
            col_letter = get_column_letter(col)
            self.calculate_col(col_letter, calculate_type, start_row)

    def auto_col_width(self, col):
        """
        自动调整列宽
        :param col: 列号
        """
        self.sheet.column_dimensions[col].auto_size = True
        for cell in self.sheet[col]:
            if cell.value:
                self.sheet.column_dimensions[col].width = max((self.sheet.column_dimensions[col].width, len(str(cell.value)) + 2))
        self.sheet.column_dimensions[col].width *= 1.2

    # def generate_pivot(self, pivot_sht, pivot_col, pivot_row, pivot_type):
    #     """
    #     生成透视表
    #     """
    #     from openpyxl.utils.dataframe import dataframe_to_rows
    #     import pandas as pd
    #     data_values = self.read_all_data()
    #     data_frame = pd.DataFrame(data_values)
    #     data_frame.columns = data_frame.iloc[0]  # 设置第一行为列名
    #     data_frame = data_frame[1:]  # 删除第一行（已设置为列名的行）
    #     pivot_table = pd.pivot_table(data_frame, values=pivot_col, index=pivot_row, aggfunc=pivot_type)  # index 行， values 列 aggfunc 计算方式：mean、sum、max、min
    #
    #     for r in dataframe_to_rows(pivot_table, index=True, header=True):
    #         pivot_sht.sheet.append(r)

    def hide_col(self, col_text:str):
        """
        隐藏指定列
        :param col_text: 列号
        """
        self.sheet.column_dimensions[col_text].hidden = True

    def hide_row(self, row_no:int):
        """
       隐藏指定行
       :param row_no: 行号
       """
        self.sheet.row_dimensions[row_no].hidden = True

    def hide_sheet(self):
        """
        隐藏当前sheet页
        """
        self.sheet.sheet_state = 'hidden'

    def hide_multi_cols(self, col_list:[str]):
        """
        批量隐藏多列
        :param col_list: 列号的列表
        """
        for col_text in col_list:
            self.hide_col(col_text)

    def hide_multi_rows(self, row_list:[int]):
        """
        批量隐藏多列
        :param row_list: 行号的列表
        """

        for row_no in row_list:
            self.hide_row(row_no)

    def show_sheet(self):
        """
        取消隐藏当前sheet页
        """
        self.sheet.sheet_state = 'visible'

    def show_multi_cols(self, col_list:[str]):
        """
        批量取消隐藏多列
        :param col_list: 列号的列表
        """
        for col_text in col_list:
            self.sheet.column_dimensions[col_text].hidden = False

    def show_multi_rows(self, row_list:[int]):
        """
        批量取消隐藏多列
        :param row_list: 行号的列表
        """

        for row_no in row_list:
            self.sheet.row_dimensions[row_no].hidden = False



